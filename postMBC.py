import glob, os, re, sys
import numpy as np
import plotCampbellData as pCD
import eigenanalysis as eigAnl
import matplotlib.pyplot as plt
import openpyxl, csv

def writeExcelData(MBC_data,OP,BladeLen,TowerHt,Pltform=False):
    StartRow=48
    StartCol_Magnitude=5
    StartCol_Phase=35
    Col_CD=3


    path2File=os.path.dirname(sys.argv[0])
    print(path2File)
    wb = openpyxl.load_workbook(path2File+'/CampbellDiagramTemplate.xlsm',read_only=False, keep_vba= True) # open the excel workbook with  macros

    for (MBC,speed) in zip(MBC_data, OP):        
        NROWS=MBC['eigSol']['NaturalFreqs_Hz'].shape[0]
        NCOLS=MBC['eigSol']['NaturalFreqs_Hz'].shape[0]

        if(NROWS>28):
            NROWS=28
            NCOLS=28

        count=0
        DescStates=[]
        if Pltform==False:
            for r in range(0,NROWS):
                if( 'platform' in MBC['DescStates'][r].lower() ):
                    count=count+1
                else:
                    DescStates.append(MBC['DescStates'][r])
        else:
            DescStates=MBC['DescStates']

        DampRatios=MBC['eigSol']['DampRatios'][:-count]
        DampFreq_Hz=MBC['eigSol']['DampedFreqs_Hz'][:-count]
        NaturalFreq_Hz=MBC['eigSol']['NaturalFreqs_Hz'][:-count]
        
        fst_tower_fa_ind = [i for i, x in enumerate(DescStates) if "tower" in x and "1st" in x and "fore-aft" in x]
        fst_tower_ss_ind = [i for i, x in enumerate(DescStates) if "tower" in x and "1st" in x and "side-to-side" in x]
        snd_tower_fa_ind = [i for i, x in enumerate(DescStates) if "tower" in x and "2nd" in x and "fore-aft" in x]
        snd_tower_ss_ind = [i for i, x in enumerate(DescStates) if "tower" in x and "2nd" in x and "side-to-side" in x]
        fst_tower_fa_ind[0]=fst_tower_fa_ind[0]+11
        fst_tower_ss_ind[0]=fst_tower_ss_ind[0]+11
        snd_tower_fa_ind[0]=snd_tower_fa_ind[0]+11
        snd_tower_ss_ind[0]=snd_tower_ss_ind[0]+11

        #print(fst_tower_ss_ind[0], fst_tower_fa_ind[0], snd_tower_ss_ind[0], snd_tower_fa_ind[0])
        fst_fw_blade_reg_ind = snd_tower_ss_ind[0]+2
        fst_fw_blade_col_ind = snd_tower_ss_ind[0]+3
        fst_fw_blade_pro_ind = snd_tower_ss_ind[0]+4
        fst_ew_blade_reg_ind = snd_tower_ss_ind[0]+5
        fst_ew_blade_col_ind = snd_tower_ss_ind[0]+6
        fst_ew_blade_pro_ind = snd_tower_ss_ind[0]+7
        snd_fw_blade_reg_ind = snd_tower_ss_ind[0]+8
        snd_fw_blade_col_ind = snd_tower_ss_ind[0]+9
        snd_fw_blade_pro_ind = snd_tower_ss_ind[0]+10

        ## get the scaling factors for the mode rows
        ScalingFactor = getScaleFactors(DescStates, TowerHt, BladeLen);

        if int(speed)%2 == 0:
            ws_CD = wb['CampbellDiagram']
            
            ws_CD.cell(row=4,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_tower_fa_ind[0])
            ws_CD.cell(row=5,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_tower_ss_ind[0])
            ws_CD.cell(row=6,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_tower_fa_ind[0])
            ws_CD.cell(row=7,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_tower_ss_ind[0])
            ws_CD.cell(row=8,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_fw_blade_reg_ind)
            ws_CD.cell(row=9,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_fw_blade_col_ind)
            ws_CD.cell(row=10,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_fw_blade_pro_ind)
            ws_CD.cell(row=11,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_ew_blade_reg_ind)
            ws_CD.cell(row=12,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_ew_blade_pro_ind)
            ws_CD.cell(row=13,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_fw_blade_reg_ind)
            ws_CD.cell(row=14,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_fw_blade_col_ind)
            ws_CD.cell(row=15,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_fw_blade_pro_ind)
            Col_CD=Col_CD+1

            
        NR=NROWS-count
        NC=NCOLS-count
        if str(int(speed))+' RPM' in wb.sheetnames:
            ws = wb[str(int(speed))+' RPM']

            # insert blade length and tower height
            ws.cell(row=4, column=5).value=BladeLen
            ws.cell(row=4, column=7).value=TowerHt
            for r in range(NR):
                ws.cell(row=11+r, column=3).value=DescStates[r]
                ws.cell(row=11+r, column=4).value=ScalingFactor[r]
                
            # insert natural frequencies, damping ratios, and damped frequencies
            for r in range(NR):
                ws.cell(row=StartRow+r, column=65).value=MBC['eigSol']['DampRatios'][r][0]
                ws.cell(row=StartRow+r, column=66).value=MBC['eigSol']['DampedFreqs_Hz'][r][0]
                ws.cell(row=StartRow+r, column=67).value=MBC['eigSol']['NaturalFreqs_Hz'][r][0]

                for r in range(NR):
                    for c in range(NC):
                        ws.cell(row=StartRow+r, column=StartCol_Magnitude+c).value=MBC['eigSol']['MagnitudeModes'][r][c]
                        ws.cell(row=StartRow+r, column=StartCol_Phase+c).value=MBC['eigSol']['PhaseModes_deg'][r][c]
        else:
            print("The worksheet '{}' does not exist in the template workbook".format(str(int(speed))+' RPM'))

        wb.save('CampbellData-v2.xlsm') #save it as a new file with macros enables


def readExcelData(fileExcel):
    xlFile = pd.ExcelFile(fileExcel)
    d = {} # dictionary to hold data from excel sheets
    for sheet in xlFile.sheet_names:
        d[f'{sheet}']= pd.read_excel(xlFile,sheet_name=sheet)
        
    frequency=d['FrequencyHz']
    OP=frequency[frequency.columns[0]].to_list()
    frequency=frequency.drop(frequency.columns[0],axis=1)
    frequency=frequency.drop(['1P','3P','6P','9P','12P'],axis=1)

    dampratio=d['DampingRatios']
    dampratio=dampratio.drop(dampratio.columns[0],axis=1)

    print(OP,frequency,dampratio)    
    pCD.plotCampbellData(OP,frequency,dampratio)
    exit()

def getFastFiles(inputFile):
    FileNames=[]

    d = {}
    with open(inputFile) as f:
        for line in f:
            if line.startswith('#') or line=='\n':
                continue
            line=line.strip() # remove whitespaces for each line  
            line=line.split('#')[0] # ignore text after # character
            key = line.split('=')[0] # split the line by = and take the first string as key
            val = line.split('=')[1] # split the line by = and take the second string as value(s)
            val = [x.strip() for x in val.split(',')]
            d[key] = val
            
    inFile=d['InputFile'][0]
    base_tempFile=os.path.splitext(os.path.basename(inFile))[0]
    ext_tempFile=os.path.splitext(os.path.basename(inFile))[1]

    if(ext_tempFile=='.xlsx'):
        if os.path.isfile(inFile):
            readExcelData(inFile) # read campbell data from excel file and plot the campbell diagram
        else:
            print('Input Excel data file does not exist!')
        exit()
    else:
        if os.path.isfile(inFile)==False:
            print('Fast template (', inFile ,') file does not exist!');
            exit()

    for file in glob.glob(base_tempFile+'-*'+ext_tempFile):
        FileNames.append(file)
    
    # sort filenames alphanumerically
    FileNames.sort(key=lambda f: int(re.sub('\D', '', f)))
    
    PltForm=False
    if "Platform" in d:
        if d['Platform'][0] == 'True':
            PltForm=True
        else:
            PltForm=False
    
    OP=d['WindSpeed_[m/s]']
    OP=[float(i) for i in OP]
    return FileNames[0:len(OP)],OP,PltForm

def getScaleFactors(DescStates, TowerLen, BladeLen):
    
    ScalingFactor = np.ones(len(DescStates))    
    
    # look at the state description strings for tower and blade
    # translational dofs:
    for i in range(len(ScalingFactor)):
        
        # look for blade dofs:
        if DescStates[i].find('blade')!=-1 or DescStates[i].find('Blade')!=-1:
           if DescStates[i].find('rotational')==-1: # make sure this isn't a rotational dof from BeamDyn
               ScalingFactor[i] = 1.0/BladeLen;
               
        # look for tower translational dofs:
        if DescStates[i].find('tower')!=-1 or DescStates[i].find('Tower')!=-1:
            ScalingFactor[i] = 1.0/TowerLen;

        # look for blade dofs:
        elif DescStates[i].find('blade')!=-1 or DescStates[i].find('Blade')!=-1:
            if DescStates[i].find('rotational')==-1: # make sure this isn't a rotational dof from BeamDyn
                ScalingFactor[i] =  1.0/BladeLen;

    return ScalingFactor

def IdentifyModes(CampbellData):
    modesDesc = {}
    modesDesc[0]=['Generator DOF (not shown)'     , 'ED Variable speed generator DOF, rad']
    modesDesc[1]=['1st Tower FA'                  , 'ED 1st tower fore-aft bending mode DOF, m']
    modesDesc[2]=['1st Tower SS'                  , 'ED 1st tower side-to-side bending mode DOF, m']
    modesDesc[3]=['1st Blade Flap (Regressive)'   , 'ED 1st flapwise bending-mode DOF of blade (sine|cosine), m', 'Blade (sine|cosine) finite element node \d rotational displacement in Y, rad']
    modesDesc[4]=['1st Blade Flap (Collective)'   , 'ED 1st flapwise bending-mode DOF of blade collective, m', 'Blade collective finite element node \d rotational displacement in Y, rad']
    modesDesc[5]=['1st Blade Flap (Progressive)'  , 'ED 1st flapwise bending-mode DOF of blade (sine|cosine), m'] # , ...# 'Blade (sine|cosine) finite element node \d rotational displacement in Y, rad']
    modesDesc[6]=['1st Blade Edge (Regressive)'   , 'ED 1st edgewise bending-mode DOF of blade (sine|cosine), m', 'Blade (sine|cosine) finite element node \d rotational displacement in X, rad']
    modesDesc[7]=['1st Blade Edge (Progressive)'  , 'ED 1st edgewise bending-mode DOF of blade (sine|cosine), m']
    modesDesc[8]=['1st Drivetrain Torsion'        , 'ED Drivetrain rotational-flexibility DOF, rad']
    modesDesc[9]=['2nd Tower FA'                  , 'ED 2nd tower fore-aft bending mode DOF, m']
    modesDesc[10]=['2nd Tower SS'                  , 'ED 2nd tower side-to-side bending mode DOF, m']
    modesDesc[11]=['2nd Blade Flap (Regressive)'   , 'ED 2nd flapwise bending-mode DOF of blade (sine|cosine), m']
    modesDesc[12]=['2nd Blade Flap (Collective)'   , 'ED 2nd flapwise bending-mode DOF of blade collective, m', 'Blade collective finite element node \d rotational displacement in Y, rad']
    modesDesc[13]=['2nd Blade Flap (Progressive)'  , 'ED 2nd flapwise bending-mode DOF of blade (sine|cosine), m'] 
    modesDesc[14]=['Nacelle Yaw (not shown)'  , 'ED Nacelle yaw DOF, rad']

    nModes = int(len(modesDesc))
    nRuns = int(len(CampbellData))
    #modesIdentified = np.zeros(nRuns,dtype=bool)
    modesIdentified={}
    modeID_table=np.zeros((nModes,nRuns))
    
    for i in range(nRuns):
        res =  [False for i in range(len(CampbellData[i]['Modes']))]
        modesIdentified[i] = res

        #print(' MODES IDENTIFIED ', modesIdentified,len(modesDesc))
    
        for modeID in range(1,len(modesDesc)): # list of modes we want to identify
            found = False;
            
            if ( len(modesDesc[modeID][1])==0 ): 
                continue;
        
            tryNumber = 0;
            
            #print(' FOUND , Trynumber ', found, tryNumber)
            while not found and tryNumber <= 2:
                m = 0;
                while not found and m < len(modesIdentified[i]):
                    m = m + 1;
                    if modesIdentified[i][m-1] or CampbellData[i]['Modes'][m-1]['NaturalFreq_Hz'] < 0.1: # already identified this mode
                        continue;

                    #print(' NF ', i, m, CampbellData[i]['Modes'][m]['NaturalFreq_Hz'])

                    if tryNumber == 0:
                        stateMax=np.argwhere((CampbellData[i]['Modes'][m-1]['StateHasMaxAtThisMode']==1))
                        #print(' FF ',i,m,len(stateMax),type(stateMax))

                        maxDesc=[CampbellData[i]['Modes'][m-1]['DescStates'][smIndx] for smIndx in stateMax.flatten()]
                        #print(' TR0 sM ',i, m , tryNumber, len(maxDesc), maxDesc)
                        #maxDesc = CampbellData[i]['Modes'][m]['DescStates'][stateMaxIndx]

                        if len(maxDesc)==0:
                            tryNumber = tryNumber + 1;
                
                    if tryNumber > 0:
                        if tryNumber < len(CampbellData[i]['Modes'][m-1]['DescStates']):
                            stateMax=np.argwhere((CampbellData[i]['Modes'][m-1]['StateHasMaxAtThisMode']==0))

                            maxDesc=[CampbellData[i]['Modes'][m-1]['DescStates'][smIndx] for smIndx in stateMax.flatten()]
                            #print(' TY1 sM ',i, m , tryNumber, len(maxDesc))
                        
                            #maxDesc = CampbellData[i]['Modes'][m]['DescStates'][~CampbellData[i]['Modes'][m]['StateHasMaxAtThisMode']]
                            #print(maxDesc)
                        else:
                            maxDesc = [];
                    
                    j = 0;
                    while not found and j < len(maxDesc):
                        j = j + 1;
                        #print(' GGG00 ',j, len(modesDesc[modeID]))
                        for iExp in range(1,len(modesDesc[modeID])):
                            #print(' GGG0 ',iExp)
                            if re.search(modesDesc[modeID][iExp],maxDesc[j-1],re.IGNORECASE)!=None:
                                modesIdentified[i][m-1] = True;
                                #print(' GGG1 ',i,j,m, modeID, iExp, tryNumber, maxDesc[j-1], len(maxDesc))
                                modeID_table[modeID,i] = m-1
                                found = True;
                                break;
                tryNumber = tryNumber + 1;

    return modeID_table,modesDesc
