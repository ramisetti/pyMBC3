#############################################################
#	Perform MBC3 Analysis and plot Campbell Diagram     #
#	Authors: Srinivasa B. Ramisetti    	            #
#	Created:   18-July-2020		  	            #
#	Revised:   11-September-2020		            #
#	E-mail: ramisettisrinivas@yahoo.com		    #
#	Web:	http://ramisetti.github.io		    #
#############################################################
#!/usr/bin/env python

import glob
import numpy as np
import re,os,sys, os.path
import pandas as pd
import plotCampbellData as pCD
import eigenanalysis as eigAnl
import matplotlib.pyplot as plt
import openpyxl, csv

def writeExcelData(MBC_data,OP,BladeLen,TowerHt,Pltform=False):
    StartRow=48
    StartCol_Magnitude=5
    StartCol_Phase=35
    Col_CD=3


    path2File=os.path.dirname(sys.argv[0])
    print(path2File)
    wb = openpyxl.load_workbook(path2File+'/CampbellDiagram_Template_v1-FFF.xlsm',read_only=False, keep_vba= True) # open the excel workbook with  macros

    for (MBC,speed) in zip(MBC_data, OP):        
        NROWS=MBC['eigSol']['NaturalFreqs_Hz'].shape[0]
        NCOLS=MBC['eigSol']['NaturalFreqs_Hz'].shape[0]

        if(NROWS>28):
            NROWS=28
            NCOLS=28

        count=0
        DescStates=[]
        if Pltform==False:
            for r in range(0,NROWS):
                if( 'platform' in MBC['DescStates'][r].lower() ):
                    count=count+1
                else:
                    DescStates.append(MBC['DescStates'][r])
        else:
            DescStates=MBC['DescStates']

        DampRatios=MBC['eigSol']['DampRatios'][:-count]
        DampFreq_Hz=MBC['eigSol']['DampedFreqs_Hz'][:-count]
        NaturalFreq_Hz=MBC['eigSol']['NaturalFreqs_Hz'][:-count]
        
        fst_tower_fa_ind = [i for i, x in enumerate(DescStates) if "tower" in x and "1st" in x and "fore-aft" in x]
        fst_tower_ss_ind = [i for i, x in enumerate(DescStates) if "tower" in x and "1st" in x and "side-to-side" in x]
        snd_tower_fa_ind = [i for i, x in enumerate(DescStates) if "tower" in x and "2nd" in x and "fore-aft" in x]
        snd_tower_ss_ind = [i for i, x in enumerate(DescStates) if "tower" in x and "2nd" in x and "side-to-side" in x]
        fst_tower_fa_ind[0]=fst_tower_fa_ind[0]+11
        fst_tower_ss_ind[0]=fst_tower_ss_ind[0]+11
        snd_tower_fa_ind[0]=snd_tower_fa_ind[0]+11
        snd_tower_ss_ind[0]=snd_tower_ss_ind[0]+11

        #print(fst_tower_ss_ind[0], fst_tower_fa_ind[0], snd_tower_ss_ind[0], snd_tower_fa_ind[0])
        fst_fw_blade_reg_ind = snd_tower_ss_ind[0]+2
        fst_fw_blade_col_ind = snd_tower_ss_ind[0]+3
        fst_fw_blade_pro_ind = snd_tower_ss_ind[0]+4
        fst_ew_blade_reg_ind = snd_tower_ss_ind[0]+5
        fst_ew_blade_col_ind = snd_tower_ss_ind[0]+6
        fst_ew_blade_pro_ind = snd_tower_ss_ind[0]+7
        snd_fw_blade_reg_ind = snd_tower_ss_ind[0]+8
        snd_fw_blade_col_ind = snd_tower_ss_ind[0]+9
        snd_fw_blade_pro_ind = snd_tower_ss_ind[0]+10

        ## get the scaling factors for the mode rows
        ScalingFactor = getScaleFactors(DescStates, TowerHt, BladeLen);

        if int(speed)%2 == 0:
            ws_CD = wb['CampbellDiagram']
            
            ws_CD.cell(row=4,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_tower_fa_ind[0])
            ws_CD.cell(row=5,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_tower_ss_ind[0])
            ws_CD.cell(row=6,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_tower_fa_ind[0])
            ws_CD.cell(row=7,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_tower_ss_ind[0])
            ws_CD.cell(row=8,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_fw_blade_reg_ind)
            ws_CD.cell(row=9,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_fw_blade_col_ind)
            ws_CD.cell(row=10,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_fw_blade_pro_ind)
            ws_CD.cell(row=11,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_ew_blade_reg_ind)
            ws_CD.cell(row=12,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(fst_ew_blade_pro_ind)
            ws_CD.cell(row=13,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_fw_blade_reg_ind)
            ws_CD.cell(row=14,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_fw_blade_col_ind)
            ws_CD.cell(row=15,column=Col_CD).value="='"+str(int(speed))+" RPM'"+'!$BO'+str(snd_fw_blade_pro_ind)
            Col_CD=Col_CD+1

            
        NR=NROWS-count
        NC=NCOLS-count
        if str(int(speed))+' RPM' in wb.sheetnames:
            ws = wb[str(int(speed))+' RPM']

            # insert blade length and tower height
            ws.cell(row=4, column=5).value=BladeLen
            ws.cell(row=4, column=7).value=TowerHt
            for r in range(NR):
                ws.cell(row=11+r, column=3).value=DescStates[r]
                ws.cell(row=11+r, column=4).value=ScalingFactor[r]
                
            # insert natural frequencies, damping ratios, and damped frequencies
            for r in range(NR):
                ws.cell(row=StartRow+r, column=65).value=MBC['eigSol']['DampRatios'][r][0]
                ws.cell(row=StartRow+r, column=66).value=MBC['eigSol']['DampedFreqs_Hz'][r][0]
                ws.cell(row=StartRow+r, column=67).value=MBC['eigSol']['NaturalFreqs_Hz'][r][0]

                for r in range(NR):
                    for c in range(NC):
                        ws.cell(row=StartRow+r, column=StartCol_Magnitude+c).value=MBC['eigSol']['MagnitudeModes'][r][c]
                        ws.cell(row=StartRow+r, column=StartCol_Phase+c).value=MBC['eigSol']['PhaseModes_deg'][r][c]
        else:
            print("The worksheet '{}' does not exist in the template workbook".format(str(int(speed))+' RPM'))

        wb.save('newfile_plf.xlsm') #save it as a new file with macros enables


def readExcelData(fileExcel):
    xlFile = pd.ExcelFile(fileExcel)
    d = {} # dictionary to hold data from excel sheets
    for sheet in xlFile.sheet_names:
        d[f'{sheet}']= pd.read_excel(xlFile,sheet_name=sheet)
        
    frequency=d['FrequencyHz']
    OP=frequency[frequency.columns[0]].to_list()
    frequency=frequency.drop(frequency.columns[0],axis=1)
    frequency=frequency.drop(['1P','3P','6P','9P','12P'],axis=1)

    dampratio=d['DampingRatios']
    dampratio=dampratio.drop(dampratio.columns[0],axis=1)

    print(OP,frequency,dampratio)    
    pCD.plotCampbellData(OP,frequency,dampratio)
    exit()

def getFastFiles(inputFile):
    FileNames=[]

    d = {}
    with open(inputFile) as f:
        for line in f:
            if line.startswith('#') or line=='\n':
                continue
            line=line.strip() # remove whitespaces for each line  
            line=line.split('#')[0] # ignore text after # character
            key = line.split('=')[0] # split the line by = and take the first string as key
            val = line.split('=')[1] # split the line by = and take the second string as value(s)
            val = [x.strip() for x in val.split(',')]
            d[key] = val
            
    inFile=d['InputFile'][0]
    base_tempFile=os.path.splitext(os.path.basename(inFile))[0]
    ext_tempFile=os.path.splitext(os.path.basename(inFile))[1]

    if(ext_tempFile=='.xlsx'):
        if os.path.isfile(inFile):
            readExcelData(inFile) # read campbell data from excel file and plot the campbell diagram
        else:
            print('Input Excel data file does not exist!')
        exit()
    else:
        if os.path.isfile(inFile)==False:
            print('Fast template (', inFile ,') file does not exist!');
            exit()

    for file in glob.glob(base_tempFile+'-*'+ext_tempFile):
        FileNames.append(file)
    
    # sort filenames alphanumerically
    FileNames.sort(key=lambda f: int(re.sub('\D', '', f)))
    
    PltForm=False
    if "Platform" in d:
        if d['Platform'][0] == 'True':
            PltForm=True
        else:
            PltForm=False
    
    OP=d['WindSpeed_[m/s]']
    OP=[float(i) for i in OP]
    return FileNames[0:len(OP)],OP,PltForm

def getScaleFactors(DescStates, TowerLen, BladeLen):
    
    ScalingFactor = np.ones(len(DescStates))    
    
    # look at the state description strings for tower and blade
    # translational dofs:
    for i in range(len(ScalingFactor)):
        
        # look for blade dofs:
        if DescStates[i].find('blade')!=-1 or DescStates[i].find('Blade')!=-1:
           if DescStates[i].find('rotational')==-1: # make sure this isn't a rotational dof from BeamDyn
               ScalingFactor[i] = 1.0/BladeLen;
               
        # look for tower translational dofs:
        if DescStates[i].find('tower')!=-1 or DescStates[i].find('Tower')!=-1:
            ScalingFactor[i] = 1.0/TowerLen;

        # look for blade dofs:
        elif DescStates[i].find('blade')!=-1 or DescStates[i].find('Blade')!=-1:
            if DescStates[i].find('rotational')==-1: # make sure this isn't a rotational dof from BeamDyn
                ScalingFactor[i] =  1.0/BladeLen;

    return ScalingFactor


def IdentifyModes(CampbellData):
    modesDesc = {}
    modesDesc[0]=['Generator DOF (not shown)'     , 'ED Variable speed generator DOF, rad']
    modesDesc[1]=['1st Tower FA'                  , 'ED 1st tower fore-aft bending mode DOF, m']
    modesDesc[2]=['1st Tower SS'                  , 'ED 1st tower side-to-side bending mode DOF, m']
    modesDesc[3]=['1st Blade Flap (Regressive)'   , 'ED 1st flapwise bending-mode DOF of blade (sine|cosine), m', 'Blade (sine|cosine) finite element node \d rotational displacement in Y, rad']
    modesDesc[4]=['1st Blade Flap (Collective)'   , 'ED 1st flapwise bending-mode DOF of blade collective, m', 'Blade collective finite element node \d rotational displacement in Y, rad']
    modesDesc[5]=['1st Blade Flap (Progressive)'  , 'ED 1st flapwise bending-mode DOF of blade (sine|cosine), m'] # , ...# 'Blade (sine|cosine) finite element node \d rotational displacement in Y, rad']
    modesDesc[6]=['1st Blade Edge (Regressive)'   , 'ED 1st edgewise bending-mode DOF of blade (sine|cosine), m', 'Blade (sine|cosine) finite element node \d rotational displacement in X, rad']
    modesDesc[7]=['1st Blade Edge (Progressive)'  , 'ED 1st edgewise bending-mode DOF of blade (sine|cosine), m']
    modesDesc[8]=['1st Drivetrain Torsion'        , 'ED Drivetrain rotational-flexibility DOF, rad']
    modesDesc[9]=['2nd Tower FA'                  , 'ED 2nd tower fore-aft bending mode DOF, m']
    modesDesc[10]=['2nd Tower SS'                  , 'ED 2nd tower side-to-side bending mode DOF, m']
    modesDesc[11]=['2nd Blade Flap (Regressive)'   , 'ED 2nd flapwise bending-mode DOF of blade (sine|cosine), m']
    modesDesc[12]=['2nd Blade Flap (Collective)'   , 'ED 2nd flapwise bending-mode DOF of blade collective, m', 'Blade collective finite element node \d rotational displacement in Y, rad']
    modesDesc[13]=['2nd Blade Flap (Progressive)'  , 'ED 2nd flapwise bending-mode DOF of blade (sine|cosine), m'] 
    modesDesc[14]=['Nacelle Yaw (not shown)'  , 'ED Nacelle yaw DOF, rad']

    nModes = int(len(modesDesc))
    nRuns = int(len(CampbellData))
    #modesIdentified = np.zeros(nRuns,dtype=bool)
    modesIdentified={}
    modeID_table=np.zeros((nModes,nRuns))
    
    for i in range(nRuns):
        res =  [False for i in range(len(CampbellData[i]['Modes']))]
        modesIdentified[i] = res

        #print(' MODES IDENTIFIED ', modesIdentified,len(modesDesc))
    
        for modeID in range(1,len(modesDesc)): # list of modes we want to identify
            found = False;
            
            if ( len(modesDesc[modeID][1])==0 ): 
                continue;
        
            tryNumber = 0;
            
            #print(' FOUND , Trynumber ', found, tryNumber)
            while not found and tryNumber <= 2:
                m = 0;
                while not found and m < len(modesIdentified[i]):
                    m = m + 1;
                    if modesIdentified[i][m-1] or CampbellData[i]['Modes'][m-1]['NaturalFreq_Hz'] < 0.1: # already identified this mode
                        continue;

                    #print(' NF ', i, m, CampbellData[i]['Modes'][m]['NaturalFreq_Hz'])

                    if tryNumber == 0:
                        stateMax=np.argwhere((CampbellData[i]['Modes'][m-1]['StateHasMaxAtThisMode']==1))
                        #print(' FF ',i,m,len(stateMax),type(stateMax))

                        maxDesc=[CampbellData[i]['Modes'][m-1]['DescStates'][smIndx] for smIndx in stateMax.flatten()]
                        #print(' TR0 sM ',i, m , tryNumber, len(maxDesc), maxDesc)
                        #maxDesc = CampbellData[i]['Modes'][m]['DescStates'][stateMaxIndx]

                        if len(maxDesc)==0:
                            tryNumber = tryNumber + 1;
                
                    if tryNumber > 0:
                        if tryNumber < len(CampbellData[i]['Modes'][m-1]['DescStates']):
                            stateMax=np.argwhere((CampbellData[i]['Modes'][m-1]['StateHasMaxAtThisMode']==0))

                            maxDesc=[CampbellData[i]['Modes'][m-1]['DescStates'][smIndx] for smIndx in stateMax.flatten()]
                            #print(' TY1 sM ',i, m , tryNumber, len(maxDesc))
                        
                            #maxDesc = CampbellData[i]['Modes'][m]['DescStates'][~CampbellData[i]['Modes'][m]['StateHasMaxAtThisMode']]
                            #print(maxDesc)
                        else:
                            maxDesc = [];
                    
                    j = 0;
                    while not found and j < len(maxDesc):
                        j = j + 1;
                        #print(' GGG00 ',j, len(modesDesc[modeID]))
                        for iExp in range(1,len(modesDesc[modeID])):
                            #print(' GGG0 ',iExp)
                            if re.search(modesDesc[modeID][iExp],maxDesc[j-1],re.IGNORECASE)!=None:
                                modesIdentified[i][m-1] = True;
                                #print(' GGG1 ',i,j,m, modeID, iExp, tryNumber, maxDesc[j-1], len(maxDesc))
                                modeID_table[modeID,i] = m-1
                                found = True;
                                break;
                tryNumber = tryNumber + 1;

    return modeID_table,modesDesc

def campbell_diagram_data(mbc_data, BladeLen, TowerLen):
    CampbellData={}
    usePercent = True;
    #
    # mbc_data.eigSol = eiganalysis(mbc_data.AvgA);
    ndof = mbc_data['ndof2'] + mbc_data['ndof1']; #size(mbc_data.AvgA,1)/2;          # number of translational states
    nModes = len(mbc_data['eigSol']['Evals'])
    #print(nModes)
    DescStates = PrettyStateDescriptions(mbc_data['DescStates'], mbc_data['ndof2'], mbc_data['performedTransformation']);

    ## store indices of max mode for state and to order natural frequencies
    #StatesMaxMode_vals = np.amax(mbc_data['eigSol']['MagnitudeModes'],axis=1); # find which mode has the maximum value for each state (max of each row before scaling)
    StatesMaxMode = np.argmax(mbc_data['eigSol']['MagnitudeModes'],axis=1); # find which mode has the maximum value for each state (max of each row before scaling)
    SortedFreqIndx = np.argsort((mbc_data['eigSol']['NaturalFreqs_Hz']).flatten(),kind="heapsort");
    #print(SortedFreqIndx)


    if BladeLen!=0 or TowerLen!=0:
        ## get the scaling factors for the mode rows
        ScalingFactor = getScaleFactors(DescStates, TowerLen, BladeLen);

        ## scale the magnitude of the modes by ScalingFactor (for consistent units)
        #  and then scale the columns so that their maximum is 1

        ModesMagnitude = np.matmul(np.diag(ScalingFactor), mbc_data['eigSol']['MagnitudeModes']); # scale the rows
        #print(ModesMagnitude)
        
        CampbellData['ScalingFactor'] = ScalingFactor;
    else:
        ModesMagnitude = mbc_data['eigSol']['MagnitudeModes'];

    if usePercent:
        scaleCol = 0.01*np.sum( ModesMagnitude,axis=0); # find the sum of the column, and multiply by 100 (divide here) to get a percentage
    else:
        scaleCol = np.amax(ModesMagnitude,axis=0); #find the maximum value in the column, so the first element has value of 1

    ModesMagnitude = np.matmul(ModesMagnitude,np.diag(1./scaleCol)) # scale the columns

    CampbellData['NaturalFreq_Hz'] = mbc_data['eigSol']['NaturalFreqs_Hz'][SortedFreqIndx]
    CampbellData['DampingRatio']   = mbc_data['eigSol']['DampRatios'][SortedFreqIndx]
    CampbellData['RotSpeed_rpm']   = mbc_data['RotSpeed_rpm']

    CampbellData['DampedFreq_Hz'] = mbc_data['eigSol']['DampedFreqs_Hz'][SortedFreqIndx]
    CampbellData['MagnitudePhase'] =  mbc_data['eigSol']['MagnitudeModes']
    CampbellData['PhaseDiff'] =  mbc_data['eigSol']['PhaseModes_deg']

    if 'WindSpeed' in mbc_data:
        CampbellData['WindSpeed']  = mbc_data['WindSpeed']

    #print(ModesMagnitude.shape)
    CampbellData['Modes']=[]

    for i in range(nModes):
        CData={}
        CData['NaturalFreq_Hz'] = mbc_data['eigSol']['NaturalFreqs_Hz'][SortedFreqIndx[i]]
        CData['DampedFreq_Hz']  = mbc_data['eigSol']['DampedFreqs_Hz'][SortedFreqIndx[i]];
        CData['DampingRatio']   = mbc_data['eigSol']['DampRatios'][SortedFreqIndx[i]];

        
        #print(np.argsort(ModesMagnitude[:,SortedFreqIndx[0]])[::-1])
        # sort indices in descending order
        sort_state = np.argsort( ModesMagnitude[:,SortedFreqIndx[i]])[::-1];

        #print(type(sort_state))
        CData['DescStates']=[DescStates[i] for i in sort_state]
        CData['MagnitudePhase']=ModesMagnitude[sort_state,SortedFreqIndx[i]];
        Phase =                mbc_data['eigSol']['PhaseModes_deg'][sort_state,SortedFreqIndx[i]];
        # if the phase is more than +/- 90 degrees different than the first
        # one (whose value == 1 or is the largest %), we'll stick a negative value on the magnitude:
        Phase = np.mod(Phase, 360);
    
        CData['PhaseDiff'] = np.mod( Phase - Phase[0], 360); # difference in range [0, 360)
        PhaseIndx = CData['PhaseDiff'] > 180;
        CData['PhaseDiff'][PhaseIndx] = CData['PhaseDiff'][PhaseIndx] - 360;   # move to range (-180, 180]
    
        if ~usePercent:
            PhaseIndx = CData['PhaseDiff'] > 90;
            CData['MagnitudePhase'][PhaseIndx] = -1*CData['MagnitudePhase'][PhaseIndx];
            CData['PhaseDiff'][PhaseIndx] = CData['PhaseDiff'][PhaseIndx] - 180;

            PhaseIndx = CData['PhaseDiff'] <= -90;
            CData['MagnitudePhase'][PhaseIndx] = -1*CData['MagnitudePhase'][PhaseIndx];
            CData['PhaseDiff'][PhaseIndx] = CData['PhaseDiff'][PhaseIndx] + 180;

        #print(CData['MagnitudePhase'])
        #print(CData['PhaseDiff'])

        CData['StateHasMaxAtThisMode'] = np.ones(ndof, dtype=bool);
        ix = (StatesMaxMode == SortedFreqIndx[i]);
        tmp=ix[sort_state]
        CData['StateHasMaxAtThisMode']=tmp
                    
        #print(CData['StateHasMaxAtThisMode'])
        #print(CData['NaturalFreq_Hz'])
        CampbellData['Modes'].append(CData)

    #print(CampbellData[0]['MagnitudePhase'])

    CampbellData['nColsPerMode'] = 5;
    CampbellData['ModesTable'] = {}

    for i in range(nModes):
        colStart = i*CampbellData['nColsPerMode'];
        CampbellData['ModesTable'][1, colStart+1 ] = 'Mode number:';
        CampbellData['ModesTable'][1, colStart+2 ] = i;

        CampbellData['ModesTable'][2, colStart+1 ] = 'Natural (undamped) frequency (Hz):';
        CampbellData['ModesTable'][2, colStart+2 ] = np.asscalar(CampbellData['Modes'][i]['NaturalFreq_Hz'])

        CampbellData['ModesTable'][3, colStart+1 ] = 'Damped frequency (Hz):';
        CampbellData['ModesTable'][3, colStart+2 ] = np.asscalar(CampbellData['Modes'][i]['DampedFreq_Hz'])

        CampbellData['ModesTable'][4, colStart+1 ] = 'Damping ratio (-):';
        CampbellData['ModesTable'][4, colStart+2 ] = np.asscalar(CampbellData['Modes'][i]['DampingRatio'])
        
        CampbellData['ModesTable'][5, colStart+1 ] = 'Mode ' + str(i) + ' state description';
        CampbellData['ModesTable'][5, colStart+2 ] = 'State has max at mode ' + str(i);
        if usePercent:
            CampbellData['ModesTable'][5, colStart+3 ] = 'Mode ' + str(i) + ' contribution (%)';
        else:
            CampbellData['ModesTable'][5, colStart+3 ] = 'Mode ' + str(i) + ' signed magnitude';

        CampbellData['ModesTable'][5, colStart+4 ] = 'Mode ' + str(i) + ' phase (deg)';

        # need to cross check these 4 lines
        CampbellData['ModesTable'][6,colStart+1] = CampbellData['Modes'][i]['DescStates'];
        CampbellData['ModesTable'][6,colStart+2] = CampbellData['Modes'][i]['StateHasMaxAtThisMode'];
        CampbellData['ModesTable'][6,colStart+3] = CampbellData['Modes'][i]['MagnitudePhase'];
        CampbellData['ModesTable'][6,colStart+4] = CampbellData['Modes'][i]['PhaseDiff'];
        #print(CampbellData['ModesTable'][6,colStart+3])

    pd.DataFrame(CampbellData['ModesTable']).to_csv('myfile.csv', header=False, index=False)
    # with open('dict.csv', 'w') as csv_file:  
    #     writer = csv.writer(csv_file)
    #     for key, value in CampbellData['ModesTable'].items():
    #         writer.writerow([key, value])
    return CampbellData
    

def PrettyStateDescriptions(DescStates, ndof2, performedTransformation):
    idx=np.array(list(range(0,ndof2))+list(range(ndof2*2+1,len(DescStates))))    
    tmpDS = [DescStates[i] for i in idx]
    
    if performedTransformation:
        key_vals=[['BD_1','Blade collective'],['BD_2','Blade cosine'],['BD_3','Blade sine'],['blade 1','blade collective'], ['blade 2','blade cosine'], ['blade 3','Blade sine '], ['PitchBearing1','Pitch bearing collective '], ['PitchBearing2','Pitch bearing cosine '], ['PitchBearing3','Pitch bearing sine ']]
        # Replace Substrings from String List 
        sub = dict(key_vals)
        for key, val in sub.items(): 
            for idx, ele in enumerate(tmpDS):
                if key in ele: 
                    tmpDS[idx] = ele.replace(key, val)

        StateDesc=tmpDS
    else:
        StateDesc = tmpDS
    
    for i in range(len( StateDesc )):
        First=re.split('\(',StateDesc[i],2)
        Last=re.split('\)',StateDesc[i],2)

        if len(First)>0 and len(Last)>0 and len(First[0]) != len(StateDesc[i]) and len( Last[-1] ) != len(StateDesc[i]):
            StateDesc[i] = (First[0]).strip() + Last[-1];
            #print(StateDesc[i])
        
    return StateDesc


def runMBC(FileNames,NLinTimes=None):
    CampbellData={}
    MBCData=[]
    HubRad=None;TipRad=None;
    BladeLen=None; TowerHt=None
    dataFound=False;
    indx=0
    
    for i in range(len(FileNames)):
        basename=os.path.splitext(os.path.basename(FileNames[i]))[0]
        with open(FileNames[i]) as f:
            datafile = f.readlines()

        if dataFound==False:
            for line in datafile:
                if 'EDFile' in line:
                    EDFile=line.split()[0].replace('"','')
                    with open(EDFile) as edfile:
                        for edline in edfile:
                            if 'HubRad' in  edline:
                                HubRad=float(edline.split()[0])
                            elif 'TipRad' in  edline:
                                TipRad=float(edline.split()[0])
                            elif 'TowerHt' in  edline:
                                TowerHt=float(edline.split()[0])
                            
        if((TipRad!=None and HubRad!=None) or TowerHt!=None):
            BladeLen=TipRad-HubRad
            dataFound=True
        if(TowerHt==None or BladeLen==None):
            print('TowerHt and BladeLen are not available!');
            sys.exit()

        #print('Tower ht ', TowerHt, 'Hub radius', HubRad, 'Tip radius ', TipRad, 'Blade length ', BladeLen)
        found=False
        for line in datafile:
            if found==False and 'NLinTimes' in line:
                NLinTimes=int(line.split()[0])
                found=True
                
        if NLinTimes<1:
            print('NLinTimes should be greater than 0!')
            sys.exit()
            
        linFileNames=[basename+'.'+format(x, 'd')+'.lin' for x in range(1,NLinTimes+1)]

        linFileFlag=True;
        for f in linFileNames:
            if(os.path.isfile(f)!=True):
                linFileFlag=False

        if(linFileFlag):
            print('Processing ', FileNames[i], ' file!', '   number of Linearization files to process ', NLinTimes)
            MBC_data,getMatData,FAST_linData=eigAnl.fx_mbc3(linFileNames)
            MBCData.append(MBC_data)
            print('Multi-Blade Coordinate transformation completed!');
            print('  ');
            CampbellData[indx]=campbell_diagram_data(MBC_data,BladeLen,TowerHt)
            indx=indx+1;

    return CampbellData, MBCData, BladeLen, TowerHt

# read wind/rotor speeds and fast linearization files from input file
FileNames,OP,PltForm=getFastFiles(sys.argv[1])
CampbellData,MBC_Data,BladeLen,TowerHt=runMBC(FileNames)
print('Preparing campbell diagram data!');

writeExcelData(MBC_Data,OP,BladeLen,TowerHt,PltForm);

modeID_table,modesDesc=IdentifyModes(CampbellData)

print(modeID_table,len(OP))
nModes=modeID_table.shape[0]
nRuns=modeID_table.shape[1]
cols=[item[0] for item in list(modesDesc.values())]

frequency=pd.DataFrame(np.nan, index=np.arange(nRuns), columns=cols)
dampratio=pd.DataFrame(np.nan, index=np.arange(nRuns), columns=cols)
print(nRuns,nModes)

FreqPlotData=np.zeros((nRuns,nModes))
DampPlotData=np.zeros((nRuns,nModes))
for i in range(nRuns):
    for modeID in range(len(modesDesc)): # list of modes we want to identify
        idx=int(modeID_table[modeID,i])
        FreqPlotData[i,modeID]=CampbellData[i]['Modes'][idx]['NaturalFreq_Hz']
        DampPlotData[i,modeID]=CampbellData[i]['Modes'][idx]['DampingRatio']
        #print(i,modeID,modesDesc[modeID][0],FreqPlotData[i,modeID])
    frequency.iloc[i,:]=FreqPlotData[i,:]
    dampratio.iloc[i,:]=DampPlotData[i,:]

for i in range(len(OP)):
    # for 15 DOFs
    frequency.index.values[i]=OP[i]
    dampratio.index.values[i]=OP[i]

# drop columns not required
frequency=frequency.drop(['Generator DOF (not shown)', 'Nacelle Yaw (not shown)'], axis = 1)
#frequency.drop(['Generator DOF (not shown)'], axis = 1) 
dampratio=dampratio.drop(['Generator DOF (not shown)', 'Nacelle Yaw (not shown)'], axis = 1)

pCD.plotCampbellData(OP,frequency,dampratio)

lenColumns=len(frequency.columns)

frequency['1P']=np.nan
frequency['3P']=np.nan
frequency['6P']=np.nan
frequency['9P']=np.nan
frequency['12P']=np.nan

for i in range(nRuns):
    # for 1P,3P,6P,9P,and 12P harmonics
    tmp=OP[i]/60.0
    frequency.iloc[i,lenColumns]=tmp
    frequency.iloc[i,lenColumns+1]=3*tmp
    frequency.iloc[i,lenColumns+2]=6*tmp
    frequency.iloc[i,lenColumns+3]=9*tmp
    frequency.iloc[i,lenColumns+4]=12*tmp

# uncomment to write excel file with transposed frequency data
#frequency.transpose().to_excel(r'CampbellData.xlsx')

# dbgwriter = pd.ExcelWriter('CampbellTable.xlsx', engine='xlsxwriter')
# dfggg = pd.DataFrame(CampbellData[0]['ModesTable'])
# dfggg.to_excel(dbgwriter)
# dbgwriter.save()

writer = pd.ExcelWriter('CampbellData.xlsx', engine='xlsxwriter')
frequency.to_excel(writer,sheet_name='FrequencyHz')
dampratio.to_excel(writer,sheet_name='DampingRatios')

#writer.save()
#exit()
# for debugging purpose
maxsize=0
for indx in range(len(CampbellData)):
    tmp=CampbellData[indx]['NaturalFreq_Hz'].shape[0]
    #print('Shape ', CampbellData[indx]['NaturalFreq_Hz'])
    if (maxsize<tmp):
        maxsize=tmp

print('Debug Info: max number of DOFs ', maxsize)
print('Len CampbellData', len(CampbellData))
#tmpFreq=np.empty([len(CampbellData),maxsize])
#tmpDamp=np.empty([len(CampbellData),maxsize])
tmpFreq=pd.DataFrame(np.nan, index=np.arange(len(CampbellData)),columns=np.arange(maxsize))
tmpDamp=pd.DataFrame(np.nan, index=np.arange(len(CampbellData)),columns=np.arange(maxsize))
#print(len(cols),tmpFreq.shape,CampbellData[0]['NaturalFreq_Hz'].shape)
for indx in range(len(CampbellData)):
    addsize=(maxsize-CampbellData[indx]['NaturalFreq_Hz'].shape[0])
    a=CampbellData[indx]['NaturalFreq_Hz']
    tmpArr=1E-10*np.ones(addsize)
    #print(addsize,tmpFreq.shape, a.shape)
    a=np.append(tmpArr,a)
    tmpFreq.iloc[indx,:]=a

    addsize=(maxsize-CampbellData[indx]['DampingRatio'].shape[0])
    tmpArr=1E-10*np.ones(addsize)
    b=CampbellData[indx]['DampingRatio']
    b=np.append(tmpArr,b)
    tmpDamp.iloc[indx,:]=b

tmpFreq.to_excel(writer,sheet_name='SortedFreqHz')
tmpDamp.to_excel(writer,sheet_name='SortedDampRatios')

# tmpFreqFile="FreqHz.txt"
# tmpDampFile="DampingRatios.txt"
# with open(tmpFreqFile, "a") as f:
#     np.savetxt(f,tmpFreq,fmt='%g', delimiter=' ', newline=os.linesep)

# with open(tmpDampFile, "a") as f:
#     np.savetxt(f,tmpDamp,fmt='%g', delimiter=' ', newline=os.linesep)
        
# print(CampbellData[indx]['NaturalFreq_Hz'])
# print(CampbellData[indx]['DampingRatio'])
# end of debugging


# save to excel file and close
writer.save() 

#End of script